"""Build and validate the XLSForm for Form HH/2026/v1.

The form is defined here in Python and assembled into .xlsx, rather than
hand-authored in Excel. Three reasons, all of which matter for this submission:

1. **It is diffable.** A binary .xlsx shows as "changed" in git and nothing more.
   A reviewer can see exactly which constraint moved between commits.
2. **The conversion is part of the pipeline.** `python build_form.py` writes the
   workbook, runs pyxform, and fails loudly if the form does not convert. There
   is no manual export step to forget.
3. **The constraint register is generated from the same source as the form**, so
   the two cannot drift. A register hand-maintained beside a form is wrong within
   a day.

Run:  python part2_q3/build_form.py

Language
--------
Interviews are conducted in Hausa; supervisory review and analysis are in
English. Every label and every constraint message a field user can see is
therefore bilingual (`::Hausa (ha)` / `::English (en)`), with Hausa the default.
Leaving a constraint message in English is listed as an automatic loss of marks,
and rightly - 38% of enumerators are not confident readers of English.

**The Hausa strings are indicative and must be reviewed by a native speaker
before deployment.** They are marked in the constraint register.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
FORM_DIR = HERE / "form"
FORM_ID = "bansara_hh_2026"
FORM_VERSION = "2026063001"          # yyyymmddrr - see the deployment plan

# ---------------------------------------------------------------------------
# Column order. `name`, `label`, `hint` etc. must be present for every row, so
# rows are written as dicts and missing keys become blank cells.
# ---------------------------------------------------------------------------
SURVEY_COLUMNS = [
    "type", "name", "label::Hausa (ha)", "label::English (en)",
    "hint::Hausa (ha)", "hint::English (en)",
    "required", "relevant", "constraint",
    "constraint_message::Hausa (ha)", "constraint_message::English (en)",
    "calculation", "choice_filter", "appearance", "parameters", "repeat_count",
    "read_only", "default", "trigger",
]

CHOICES_COLUMNS = [
    "list_name", "name", "label::Hausa (ha)", "label::English (en)",
]


def survey_rows() -> list[dict]:
    """The form, in document order. Question numbers follow the paper form."""
    R: list[dict] = []

    def row(**kwargs):
        R.append(kwargs)

    # ---------------------------------------------------------------- metadata
    # Collected automatically. Several exist purely for back-office quality
    # assurance and are named as such in the fabrication-detection note.
    row(type="start", name="start_time")
    row(type="end", name="end_time")
    row(type="today", name="today_date")
    row(type="deviceid", name="device_id")
    row(type="audit", name="audit",
        parameters="track-changes=true identify-user=true "
                   "track-changes-reasons=on-form-edit")

    # ============================================================ SECTION 0
    # Enumerator identification. Not on the paper form as a gate, but the paper
    # form has no way to prevent one enumerator entering another's code.
    row(type="begin_group", name="s0_login",
        **{"label::Hausa (ha)": "Shiga", "label::English (en)": "Sign in"})

    row(type="select_one_from_file staff_roster.csv", name="enumerator_code",
        **{"label::Hausa (ha)": "Lambar mai tambaya",
           "label::English (en)": "Enumerator code (1.08)"},
        required="yes", appearance="autocomplete")

    row(type="calculate", name="enum_team",
        calculation="instance('staff_roster')/root/item[name=${enumerator_code}]/team_code")
    row(type="calculate", name="enum_role",
        calculation="instance('staff_roster')/root/item[name=${enumerator_code}]/role")
    row(type="calculate", name="enum_lga",
        calculation="instance('staff_roster')/root/item[name=${enumerator_code}]/assigned_lga")
    row(type="calculate", name="enum_pin",
        calculation="instance('staff_roster')/root/item[name=${enumerator_code}]/pin")

    row(type="text", name="pin_entered",
        **{"label::Hausa (ha)": "Shigar da PIN dinka",
           "label::English (en)": "Enter your 4-digit PIN"},
        required="yes", appearance="numbers masked",
        constraint="string-length(.) = 4 and . = ${enum_pin}",
        **{"constraint_message::Hausa (ha)":
           "PIN din bai dace ba. Sake gwadawa.",
           "constraint_message::English (en)":
           "PIN does not match this enumerator code."})

    row(type="note", name="team_note",
        **{"label::Hausa (ha)": "Ƙungiya: ${enum_team}  |  LGA: ${enum_lga}",
           "label::English (en)": "Team: ${enum_team}  |  LGA: ${enum_lga}"})
    row(type="end_group", name="s0_login_end")

    # ============================================================ SECTION 1
    row(type="begin_group", name="s1_identification",
        **{"label::Hausa (ha)": "Sashe 1: Bayanin gida",
           "label::English (en)": "Section 1: Household identification"})

    # 1.01 State - single value on the paper form, so it is a calculate, not a
    # question. Asking a question with one possible answer wastes field time.
    row(type="calculate", name="q1_01_state", calculation="'BAN'")

    # 1.02-1.04 cascading selects from external media. See the external-data
    # note: 2,524 settlements cannot go in a choices worksheet.
    row(type="select_one_from_file lgas.csv", name="q1_02_lga",
        **{"label::Hausa (ha)": "Karamar hukuma",
           "label::English (en)": "1.02 Local Government Area"},
        required="yes",
        constraint=". = ${enum_lga} or ${enum_role} != 'Enumerator'",
        **{"constraint_message::Hausa (ha)":
           "Ba a ba ka wannan karamar hukuma ba.",
           "constraint_message::English (en)":
           "This LGA is not the one assigned to you. Check with your supervisor."})

    row(type="select_one_from_file wards.csv", name="q1_03_ward",
        **{"label::Hausa (ha)": "Gunduma",
           "label::English (en)": "1.03 Ward"},
        required="yes", choice_filter="lga_code=${q1_02_lga}",
        appearance="autocomplete")

    row(type="select_one_from_file settlements.csv", name="q1_04_settlement",
        **{"label::Hausa (ha)": "Ƙauye / unguwa",
           "label::English (en)": "1.04 Settlement"},
        required="yes", choice_filter="ward_code=${q1_03_ward}",
        appearance="autocomplete")

    row(type="select_one yes_no", name="q1_05_alt_name_yn",
        **{"label::Hausa (ha)": "Ana kiran wannan wuri da wani suna daban?",
           "label::English (en)":
           "1.05 Is the settlement known locally by a different name?"},
        required="yes")
    row(type="text", name="q1_05_alt_name",
        **{"label::Hausa (ha)": "Rubuta sunan da ake amfani da shi a nan",
           "label::English (en)": "1.05 Local name"},
        relevant="${q1_05_alt_name_yn} = '1'", required="yes")

    row(type="integer", name="q1_06_structure",
        **{"label::Hausa (ha)": "Lambar gini da aka rubuta a gidan",
           "label::English (en)": "1.06 Structure number painted on the dwelling"},
        required="yes", constraint=". >= 1 and . <= 999",
        **{"constraint_message::Hausa (ha)": "Lamba tsakanin 1 zuwa 999.",
           "constraint_message::English (en)": "Enter a number between 1 and 999."})

    row(type="integer", name="q1_07_hh_serial",
        **{"label::Hausa (ha)": "Lambar gida a cikin ƙauyen",
           "label::English (en)": "1.07 Household serial number within the settlement"},
        required="yes", constraint=". >= 1 and . <= 999",
        **{"constraint_message::Hausa (ha)": "Lamba tsakanin 1 zuwa 999.",
           "constraint_message::English (en)": "Enter a number between 1 and 999."})

    # 1.10 date of visit. Constrained to the ethics-approved window, not the
    # 14-day expectation - see the constraint register.
    row(type="date", name="q1_10_visit_date",
        **{"label::Hausa (ha)": "Ranar ziyara",
           "label::English (en)": "1.10 Date of visit"},
        required="yes", default="today()",
        constraint=". >= date('2026-06-01') and . <= date('2026-06-30')",
        **{"constraint_message::Hausa (ha)":
           "Ranar dole ta kasance tsakanin 1 zuwa 30 ga Yuni 2026.",
           "constraint_message::English (en)":
           "Date must fall within the fieldwork window, 1-30 June 2026."})

    row(type="geopoint", name="q1_11_gps",
        **{"label::Hausa (ha)": "Ɗauki wurin GPS a ƙofar gidan",
           "label::English (en)": "1.11 GPS reading at the entrance to the dwelling"},
        required="yes")

    row(type="select_one yes_no_dk", name="q1_12_prev_round",
        **{"label::Hausa (ha)": "An ziyarci wannan gida a watan Oktoba 2025?",
           "label::English (en)":
           "1.12 Was this household visited during the October 2025 round?"},
        required="yes")

    row(type="select_one_from_file previous_round_households.csv",
        name="q1_13_prev_id",
        **{"label::Hausa (ha)": "Lambar gidan da aka ba shi a 2025",
           "label::English (en)":
           "1.13 Household identifier allocated in the October 2025 round"},
        relevant="${q1_12_prev_round} = '1'", required="yes",
        choice_filter="settlement_id=${q1_04_settlement}",
        appearance="autocomplete")

    row(type="select_one result_of_visit", name="q1_14_result",
        **{"label::Hausa (ha)": "Sakamakon ziyarar",
           "label::English (en)": "1.14 Result of visit"},
        required="yes")

    row(type="note", name="q1_14_stop_note",
        **{"label::Hausa (ha)":
           "Kada ka ci gaba. Sa hannu a 7.03 sannan ka mika wa mai kula da kai.",
           "label::English (en)":
           "Do not complete any further section. Sign at 7.03 and hand the form "
           "to your supervisor."},
        relevant="${q1_14_result} != '1'")
    row(type="end_group", name="s1_end")

    # ============================================================ SECTION 2
    row(type="begin_group", name="s2_consent",
        **{"label::Hausa (ha)": "Sashe 2: Yarda",
           "label::English (en)": "Section 2: Consent"},
        relevant="${q1_14_result} = '1'")

    # B3 - the only hard block in the form. Consent recorded after an unread
    # consent statement is not informed consent, and this form collects
    # biological specimens from children.
    row(type="select_one yes_no", name="q2_01_statement_read",
        **{"label::Hausa (ha)":
           "An karanta sanarwar yarda gaba ɗaya ga mai amsa?",
           "label::English (en)":
           "2.01 Consent statement read aloud to the respondent in full?"},
        required="yes", constraint=". = '1'",
        **{"constraint_message::Hausa (ha)":
           "Dole ne a karanta sanarwar yarda gaba ɗaya kafin a ci gaba.",
           "constraint_message::English (en)":
           "The consent statement must be read in full before continuing. "
           "Read it now, then record Yes."})

    row(type="select_one consent", name="q2_02_consent",
        **{"label::Hausa (ha)": "Mai amsa ya yarda da tambayoyin?",
           "label::English (en)":
           "2.02 Does the respondent consent to the household interview?"},
        required="yes")

    row(type="select_one relationship", name="q2_03_relationship",
        **{"label::Hausa (ha)": "Alaƙar mai amsa da shugaban gida",
           "label::English (en)":
           "2.03 Relationship of the respondent to the head of household"},
        required="yes", relevant="${q2_02_consent} = '1'")
    row(type="end_group", name="s2_end")

    # ============================================================ SECTION 3
    row(type="begin_group", name="s3_roster",
        **{"label::Hausa (ha)": "Sashe 3: Jerin mazauna gida",
           "label::English (en)": "Section 3: Household roster"},
        relevant="${q1_14_result} = '1' and ${q2_02_consent} = '1'")

    row(type="integer", name="q3_01_hh_size",
        **{"label::Hausa (ha)": "Mutane nawa ne ke zama a wannan gida?",
           "label::English (en)":
           "3.01 How many people usually live in this household?"},
        required="yes", constraint=". >= 1 and . <= 40",
        **{"constraint_message::Hausa (ha)":
           "Adadi tsakanin 1 zuwa 40. Idan ya fi haka, sanar da mai kula da kai.",
           "constraint_message::English (en)":
           "Enter between 1 and 40. If larger, notify your supervisor."})

    # A4 - an unbounded repeat, so the paper form's 12-row ceiling disappears
    # and the 3.01-vs-roster check becomes meaningful rather than an artefact.
    row(type="begin_repeat", name="roster",
        **{"label::Hausa (ha)": "Mazaunin gida", "label::English (en)": "Resident"},
        repeat_count="${q3_01_hh_size}")

    row(type="calculate", name="line_no", calculation="position(..)")
    row(type="note", name="line_note",
        **{"label::Hausa (ha)": "Layi ${line_no}",
           "label::English (en)": "Line ${line_no}"})

    row(type="text", name="r_initials",
        **{"label::Hausa (ha)": "Baƙaƙen farko na suna",
           "label::English (en)": "(2) Initials"},
        required="yes",
        **{"hint::Hausa (ha)": "Baƙaƙen farko kawai, ba cikakken suna ba.",
           "hint::English (en)": "Initials only - do not record the full name."})

    row(type="select_one relationship", name="r_relationship",
        **{"label::Hausa (ha)": "Alaƙa da shugaban gida",
           "label::English (en)": "(3) Relationship to head"}, required="yes")

    row(type="select_one sex", name="r_sex",
        **{"label::Hausa (ha)": "Jinsi", "label::English (en)": "(4) Sex"},
        required="yes")

    row(type="select_one age_unit", name="r_age_unit",
        **{"label::Hausa (ha)": "Shekaru ko watanni?",
           "label::English (en)": "Age recorded in years or months?"},
        required="yes",
        **{"hint::Hausa (ha)": "Watanni ga yara ƙasa da shekara 5.",
           "hint::English (en)": "Months for children under 5 years."})

    row(type="integer", name="r_age_years",
        **{"label::Hausa (ha)": "Shekaru cikakku",
           "label::English (en)": "(5) Age in completed years"},
        relevant="${r_age_unit} = 'years'", required="yes",
        constraint=". >= 5 and . <= 120",
        **{"constraint_message::Hausa (ha)":
           "Shekaru 5 zuwa 120. Ga yara ƙasa da 5, yi amfani da watanni.",
           "constraint_message::English (en)":
           "5 to 120 years. For children under 5, record months instead."})

    row(type="integer", name="r_age_months",
        **{"label::Hausa (ha)": "Watanni cikakku",
           "label::English (en)": "(6) Age in completed months"},
        relevant="${r_age_unit} = 'months'", required="yes",
        constraint=". >= 0 and . <= 59",
        **{"constraint_message::Hausa (ha)":
           "Watanni 0 zuwa 59. Idan ya kai watanni 60, yi amfani da shekaru.",
           "constraint_message::English (en)":
           "0 to 59 months. At 60 months and above, record age in years."})

    # A1 - eligibility is DERIVED, never transcribed from an office-use column.
    row(type="calculate", name="r_eligible",
        calculation="if(${r_age_unit} = 'months' and ${r_age_months} >= 9 "
                    "and ${r_age_months} <= 59, 1, 0)")
    row(type="end_repeat", name="roster_end")

    # F4 - cross-question consistency the paper form leaves to a clerk.
    row(type="calculate", name="roster_count", calculation="count(${roster})")
    row(type="calculate", name="q3_02_eligible",
        calculation="sum(${r_eligible})")

    row(type="note", name="roster_mismatch_note",
        **{"label::Hausa (ha)":
           "⚠ Ka ce mutane ${q3_01_hh_size} ne, amma ka rubuta ${roster_count}. "
           "Duba kafin ka ci gaba.",
           "label::English (en)":
           "⚠ You recorded ${q3_01_hh_size} usual residents but listed "
           "${roster_count}. Check the roster before continuing."},
        relevant="${roster_count} != ${q3_01_hh_size}")

    row(type="note", name="q3_02_note",
        **{"label::Hausa (ha)":
           "Yara masu watanni 9-59 a wannan gida: ${q3_02_eligible}",
           "label::English (en)":
           "3.02 Children aged 9-59 completed months in this household: "
           "${q3_02_eligible}"})
    row(type="end_group", name="s3_end")

    # ================================================== SECTIONS 4 AND 5
    # B2 - Sections 4 and 5 sit inside ONE per-child repeat. On paper, 5.01
    # sends an under-12-month child to Section 6, abandoning every remaining
    # child in the household. Here "skip" means "end this child's iteration".
    row(type="begin_repeat", name="child",
        **{"label::Hausa (ha)": "Yaro", "label::English (en)": "Child"},
        repeat_count="${q3_02_eligible}",
        relevant="${q1_14_result} = '1' and ${q2_02_consent} = '1' "
                 "and ${q3_02_eligible} > 0")

    row(type="calculate", name="child_index", calculation="position(..)")

    # 4.01 is kept as the paper form asks it - the enumerator gives the roster
    # line number - but it is now validated against the roster itself. On paper
    # nothing prevents pointing a child module at an adult, or at a line that
    # does not exist. indexed-repeat() reaches into the roster by line number.
    row(type="integer", name="q4_01_line",
        **{"label::Hausa (ha)": "Layin yaron a jerin mazauna",
           "label::English (en)":
           "4.01 Line number of this child in the Section 3 roster"},
        required="yes",
        constraint=". >= 1 and . <= ${roster_count} and "
                   "indexed-repeat(${r_eligible}, ${roster}, .) = 1",
        **{"constraint_message::Hausa (ha)":
           "Wannan layin ba yaro mai watanni 9-59 ba ne. Duba jerin.",
           "constraint_message::English (en)":
           "That line is not a child aged 9-59 completed months. Check the roster."})

    row(type="calculate", name="q4_02_initials",
        calculation="indexed-repeat(${r_initials}, ${roster}, ${q4_01_line})")
    row(type="calculate", name="q4_03_age_months",
        calculation="indexed-repeat(${r_age_months}, ${roster}, ${q4_01_line})")

    row(type="note", name="child_age_note",
        **{"label::Hausa (ha)": "Watanni: ${q4_03_age_months}",
           "label::English (en)":
           "4.03 Age in completed months (from roster): ${q4_03_age_months}"})

    row(type="select_one sex", name="q4_04_sex",
        **{"label::Hausa (ha)": "Jinsin yaron", "label::English (en)": "4.04 Sex of the child"},
        required="yes")

    # D3 - measurement and measurement STATUS are separate fields. The paper
    # form uses 99 for "not measured", and 99 cm is an ordinary height for a
    # three-year-old. No sentinel is ever stored inside a measurement.
    row(type="select_one measured", name="q4_05_weight_status",
        **{"label::Hausa (ha)": "An auna nauyin yaron?",
           "label::English (en)": "4.05 Was the child weighed?"}, required="yes")
    row(type="decimal", name="q4_05_weight_kg",
        **{"label::Hausa (ha)": "Nauyi (kg)", "label::English (en)": "4.05 Weight in kg"},
        relevant="${q4_05_weight_status} = 'measured'", required="yes",
        constraint=". >= 2.0 and . <= 30.0",
        **{"constraint_message::Hausa (ha)": "Nauyi tsakanin 2.0 zuwa 30.0 kg.",
           "constraint_message::English (en)":
           "Weight must be between 2.0 and 30.0 kg for a child aged 9-59 months."})

    row(type="select_one measured", name="q4_06_height_status",
        **{"label::Hausa (ha)": "An auna tsayin yaron?",
           "label::English (en)": "4.06 Was the child measured?"}, required="yes")
    row(type="decimal", name="q4_06_height_cm",
        **{"label::Hausa (ha)": "Tsayi (cm)", "label::English (en)": "4.06 Length or height in cm"},
        relevant="${q4_06_height_status} = 'measured'", required="yes",
        constraint=". >= 45.0 and . <= 130.0",
        **{"constraint_message::Hausa (ha)": "Tsayi tsakanin 45.0 zuwa 130.0 cm.",
           "constraint_message::English (en)":
           "Height must be between 45.0 and 130.0 cm for a child aged 9-59 months."})

    # WHO measurement convention: recumbent under 24 months, standing at 24+.
    row(type="select_one measure_position", name="q4_07_position",
        **{"label::Hausa (ha)": "Yadda aka auna yaron",
           "label::English (en)": "4.07 Position in which the child was measured"},
        required="yes", relevant="${q4_06_height_status} = 'measured'")
    row(type="note", name="q4_07_position_warn",
        **{"label::Hausa (ha)":
           "⚠ Ana auna yara ƙasa da watanni 24 a kwance, na watanni 24 sama a tsaye.",
           "label::English (en)":
           "⚠ Convention is recumbent below 24 months and standing at 24 months "
           "and above. Confirm this was intended."},
        relevant="(${q4_03_age_months} < 24 and ${q4_07_position} = '2') or "
                 "(${q4_03_age_months} >= 24 and ${q4_07_position} = '1')")

    row(type="select_one card_seen", name="q4_08_card",
        **{"label::Hausa (ha)": "Za a iya ganin katin rigakafi na yaron?",
           "label::English (en)":
           "4.08 May I see the child's vaccination card or health record?"},
        required="yes")
    # A2 - the paper question asks for a three-way distinction its coding cannot
    # hold. Added as an optional follow-up; flagged in the codebook.
    row(type="select_one document_type", name="q4_08a_doc_type",
        **{"label::Hausa (ha)": "Wane irin takarda aka gani?",
           "label::English (en)":
           "4.08a Which document was seen? (addition - not on the paper form)"},
        relevant="${q4_08_card} = '1'")

    row(type="select_one yes_no", name="q4_09_measles_card",
        **{"label::Hausa (ha)": "An rubuta allurar kyanda a katin?",
           "label::English (en)":
           "4.09 Copy from the card: is a measles dose recorded?"},
        required="yes", relevant="${q4_08_card} = '1'")

    row(type="select_one yes_no_dk", name="q4_10_measles_recall",
        **{"label::Hausa (ha)": "An taɓa yi wa yaron allurar kyanda?",
           "label::English (en)":
           "4.10 Has this child ever received a measles vaccination?"},
        required="yes", relevant="${q4_08_card} = '2'")

    row(type="select_one yes_no_dk", name="q4_11_diarrhoea",
        **{"label::Hausa (ha)": "Yaron ya yi gudawa a cikin kwana 14 da suka wuce?",
           "label::English (en)":
           "4.11 Has this child had diarrhoea in the past 14 days?"}, required="yes")

    row(type="select_one yes_no_dk", name="q4_12_antibiotic",
        **{"label::Hausa (ha)":
           "Yaron ya sha maganin ƙwayoyin cuta a cikin kwana 30 da suka wuce?",
           "label::English (en)":
           "4.12 Has this child taken any antibiotic medicine in the past 30 days?"},
        required="yes")

    # C1 - the paper form keeps only the most recent antibiotic, in an AMR
    # survey. Escalated rather than changed; this single addition lets analysis
    # know when the recorded code is incomplete.
    row(type="select_one yes_no_dk", name="q4_12a_more_than_one",
        **{"label::Hausa (ha)": "An sha fiye da magani ɗaya?",
           "label::English (en)":
           "4.12a Was more than one antibiotic taken? (addition - see defect C1)"},
        relevant="${q4_12_antibiotic} = '1'", required="yes")

    # E1 - PLACEHOLDER codelist. The medicine list referenced by the paper form
    # is not in the data pack. Values are WHO ATC codes, which cannot be
    # mistaken for the two-digit local codes the paper form expects.
    row(type="note", name="q4_13_placeholder_warning",
        **{"label::Hausa (ha)":
           "⚠ JERIN MAGUNGUNA NA GWAJI NE - BA A YI AMFANI DA SHI A AIKI BA.",
           "label::English (en)":
           "⚠ PLACEHOLDER MEDICINE LIST - NOT FOR DEPLOYMENT. The codelist "
           "referenced by the paper form was not supplied. See defect E1."},
        relevant="${q4_12_antibiotic} = '1'")
    row(type="select_one_from_file medicines.csv", name="q4_13_medicine",
        **{"label::Hausa (ha)": "Wane magani aka sha?",
           "label::English (en)": "4.13 Which antibiotic was taken?"},
        relevant="${q4_12_antibiotic} = '1'", required="yes",
        appearance="autocomplete")
    row(type="text", name="q4_14_medicine_other",
        **{"label::Hausa (ha)": "Rubuta sunan maganin kamar yadda aka faɗa",
           "label::English (en)": "4.14 Write the name of the medicine as reported"},
        relevant="${q4_13_medicine} = 'OTHER96'", required="yes")

    row(type="select_one yes_no_dk", name="q4_15_no_prescription",
        **{"label::Hausa (ha)": "An sami maganin ba tare da takardar likita ba?",
           "label::English (en)":
           "4.15 Was the medicine obtained without a prescription from a health worker?"},
        relevant="${q4_12_antibiotic} = '1'", required="yes")

    row(type="select_one photo_taken", name="q4_16_photo_status",
        **{"label::Hausa (ha)": "An ɗauki hoton kunshin maganin?",
           "label::English (en)":
           "4.16 Was a photograph of the medicine packaging taken?"},
        relevant="${q4_12_antibiotic} = '1'", required="yes")
    row(type="image", name="q4_16_photo",
        **{"label::Hausa (ha)": "Ɗauki hoton kunshin maganin",
           "label::English (en)": "4.16 Photograph of the medicine packaging"},
        relevant="${q4_16_photo_status} = '1'", required="yes",
        parameters="max-pixels=1024")

    # --------------------------------------------------- Section 5, per child
    row(type="begin_group", name="s5_specimen",
        **{"label::Hausa (ha)": "Sashe 5: Tattara samfuri",
           "label::English (en)": "Section 5: Specimen collection"})

    # A3 - calculated, not asked. The paper instruction says "every eligible
    # child" while 5.01 restricts to 12 months and over; the filter governs.
    row(type="calculate", name="q5_01_specimen_eligible",
        calculation="if(${q4_03_age_months} >= 12, 1, 0)")
    row(type="note", name="q5_01_note",
        **{"label::Hausa (ha)": "Yaron bai kai watanni 12 ba - ba a neman samfuri.",
           "label::English (en)":
           "5.01 Child is under 12 completed months - no specimen is sought."},
        relevant="${q5_01_specimen_eligible} = 0")

    # B1 - the missing skip. 5.02 had no skip instruction at all on paper.
    row(type="select_one yes_no", name="q5_02_specimen_obtained",
        **{"label::Hausa (ha)": "An sami samfurin bayan gida daga yaron?",
           "label::English (en)":
           "5.02 Was a stool specimen obtained from this child?"},
        required="yes", relevant="${q5_01_specimen_eligible} = 1")

    row(type="text", name="q5_03_label_serial",
        **{"label::Hausa (ha)": "Lambar lakabin samfuri (lamba 6)",
           "label::English (en)": "5.03 Specimen label serial (6 digits, after BSN)"},
        relevant="${q5_02_specimen_obtained} = '1'", required="yes",
        appearance="numbers",
        constraint="regex(., '^[0-9]{6}$') and "
                   "number(.) >= number(instance('specimen_label_allocation')/root/"
                   "item[team_code=${enum_team}]/range_start) and "
                   "number(.) <= number(instance('specimen_label_allocation')/root/"
                   "item[team_code=${enum_team}]/range_end)",
        **{"constraint_message::Hausa (ha)":
           "Lambar ba ta cikin kewayon ƙungiyar ka ba. Duba lakabin.",
           "constraint_message::English (en)":
           "Serial is not within the range allocated to your team. Check the label."})

    row(type="calculate", name="check_digit_expected",
        calculation=(
            "if("
            "(7 * number(substr(${q5_03_label_serial},0,1)) + "
            "6 * number(substr(${q5_03_label_serial},1,2)) + "
            "5 * number(substr(${q5_03_label_serial},2,3)) + "
            "4 * number(substr(${q5_03_label_serial},3,4)) + "
            "3 * number(substr(${q5_03_label_serial},4,5)) + "
            "2 * number(substr(${q5_03_label_serial},5,6))) mod 11 = 10, "
            "'X', "
            "string((7 * number(substr(${q5_03_label_serial},0,1)) + "
            "6 * number(substr(${q5_03_label_serial},1,2)) + "
            "5 * number(substr(${q5_03_label_serial},2,3)) + "
            "4 * number(substr(${q5_03_label_serial},3,4)) + "
            "3 * number(substr(${q5_03_label_serial},4,5)) + "
            "2 * number(substr(${q5_03_label_serial},5,6))) mod 11))"))

    # E3 - the check character can be X, which a digit box cannot hold.
    row(type="text", name="q5_03_check_digit",
        **{"label::Hausa (ha)": "Lambar tantancewa (bayan layin)",
           "label::English (en)": "5.03 Check character (after the hyphen)"},
        relevant="${q5_02_specimen_obtained} = '1'", required="yes",
        constraint="translate(., 'x', 'X') = ${check_digit_expected}",
        **{"constraint_message::Hausa (ha)":
           "Lambar tantancewa ba ta dace ba. Sake duba lambobin lakabin.",
           "constraint_message::English (en)":
           "Check character does not match the serial. Re-read the label - two "
           "digits may have been swapped."})

    row(type="calculate", name="specimen_label_full",
        calculation="concat('BSN', ${q5_03_label_serial}, '-', "
                    "translate(${q5_03_check_digit}, 'x', 'X'))")

    row(type="time", name="q5_04_coldbox_time",
        **{"label::Hausa (ha)": "Lokacin da aka sanya samfurin cikin akwatin sanyi",
           "label::English (en)":
           "5.04 Time the specimen was placed in the cold box"},
        relevant="${q5_02_specimen_obtained} = '1'", required="yes")

    # C3 - the paper field is one digit and one decimal, so 0.0-9.9, and cannot
    # record a cold-chain failure at all. Widened; 2-8 warns rather than blocks.
    row(type="decimal", name="q5_05_coldbox_temp",
        **{"label::Hausa (ha)": "Zafin akwatin sanyi (°C)",
           "label::English (en)": "5.05 Temperature shown on the cold box thermometer"},
        relevant="${q5_02_specimen_obtained} = '1'", required="yes",
        constraint=". >= -20.0 and . <= 40.0",
        **{"constraint_message::Hausa (ha)": "Zafi tsakanin -20.0 zuwa 40.0 °C.",
           "constraint_message::English (en)":
           "Temperature must be between -20.0 and 40.0 °C."})
    row(type="note", name="q5_05_temp_warn",
        **{"label::Hausa (ha)":
           "⚠ Zafin ya fita daga 2-8 °C. Sanar da mai kula da kai yanzu.",
           "label::English (en)":
           "⚠ Temperature is outside 2-8 °C. Report to your supervisor now - "
           "the cold chain may have failed."},
        relevant="${q5_02_specimen_obtained} = '1' and "
                 "(${q5_05_coldbox_temp} < 2 or ${q5_05_coldbox_temp} > 8)")

    row(type="select_one no_specimen_reason", name="q5_06_reason",
        **{"label::Hausa (ha)": "Dalilin da ya sa ba a sami samfuri ba",
           "label::English (en)": "5.06 Reason no specimen was obtained"},
        relevant="${q5_02_specimen_obtained} = '2'", required="yes")
    row(type="text", name="q5_07_reason_other",
        **{"label::Hausa (ha)": "Bayyana dalilin",
           "label::English (en)": "5.07 Specify"},
        relevant="${q5_06_reason} = '96'", required="yes")
    row(type="end_group", name="s5_end")
    row(type="end_repeat", name="child_end")

    # ============================================================ SECTION 6
    row(type="begin_group", name="s6_environment",
        **{"label::Hausa (ha)": "Sashe 6: Yanayin gida",
           "label::English (en)": "Section 6: Household environment"},
        relevant="${q1_14_result} = '1' and ${q2_02_consent} = '1'")

    # D2 - the paper list runs to 11 with 9 = Rainwater, colliding with the
    # single-digit no-answer sentinel. Values are stored as w01..w11 so no
    # substantive value can ever equal a sentinel.
    row(type="select_one water_source", name="q6_01_water",
        **{"label::Hausa (ha)": "Babbar hanyar samun ruwan sha ta wannan gida?",
           "label::English (en)":
           "6.01 What is the main source of drinking water for this household?"},
        required="yes")

    # D1 - same mechanism. 9 = "No facility or bush" collided with "no answer".
    row(type="select_one toilet", name="q6_02_toilet",
        **{"label::Hausa (ha)": "Wane irin bandaki wannan gida ke amfani da shi?",
           "label::English (en)":
           "6.02 What kind of toilet facility do members of this household use?"},
        required="yes")

    row(type="select_one yes_no", name="q6_03_livestock",
        **{"label::Hausa (ha)": "Ana kiwon dabbobi ko kaji a cikin gidan?",
           "label::English (en)":
           "6.03 Does this household keep poultry or livestock inside the compound?"},
        required="yes")
    row(type="select_one yes_no_dk", name="q6_04_animal_antibiotic",
        **{"label::Hausa (ha)":
           "An ba dabbobin maganin ƙwayoyin cuta cikin watanni 12 da suka wuce?",
           "label::English (en)":
           "6.04 Have any antibiotic medicines been given to these animals in the "
           "past 12 months?"},
        relevant="${q6_03_livestock} = '1'", required="yes")

    row(type="select_one handwashing", name="q6_05_handwashing",
        **{"label::Hausa (ha)":
           "Duba: akwai wurin wanke hannu da sabulu da ruwa?",
           "label::English (en)":
           "6.05 Observe: is there a handwashing station with both soap and water?"},
        required="yes")

    row(type="select_one yes_no_dk", name="q6_06_hh_diarrhoea",
        **{"label::Hausa (ha)":
           "Wani a wannan gida ya yi gudawa cikin makonni biyu da suka wuce?",
           "label::English (en)":
           "6.06 Has any member of this household had diarrhoea in the past two weeks?"},
        required="yes")

    # C2 - "None of these" was selectable alongside owned assets on paper.
    row(type="select_multiple assets", name="q6_07_assets",
        **{"label::Hausa (ha)": "Waɗanne daga cikin waɗannan wannan gida ke da su?",
           "label::English (en)":
           "6.07 Which of the following does this household own?"},
        required="yes",
        constraint="not(selected(., 'H') and count-selected(.) > 1)",
        **{"constraint_message::Hausa (ha)":
           "Ba za a iya zaɓar 'Babu ɗaya daga cikin waɗannan' tare da wasu ba.",
           "constraint_message::English (en)":
           "'None of these' cannot be selected together with any other item."})
    row(type="end_group", name="s6_end")

    # ============================================================ SECTION 7
    row(type="begin_group", name="s7_closeout",
        **{"label::Hausa (ha)": "Sashe 7: Kammalawa",
           "label::English (en)": "Section 7: Close-out and supervisor review"})

    row(type="calculate", name="interview_duration_min",
        calculation="round((decimal-date-time(${end_time}) - "
                    "decimal-date-time(${start_time})) * 1440, 1)")

    row(type="text", name="q7_02_observation",
        **{"label::Hausa (ha)": "Duk wani abin lura da zai taimaka wa ofis",
           "label::English (en)":
           "7.02 Record any observation that may help the office interpret this form"},
        appearance="multiline")

    row(type="select_one_from_file staff_roster.csv", name="q7_04_supervisor",
        **{"label::Hausa (ha)": "Lambar mai kula",
           "label::English (en)": "7.04 Supervisor code"},
        choice_filter="role='Team supervisor'", appearance="autocomplete")

    row(type="select_one supervisor_decision", name="q7_05_supervisor_decision",
        **{"label::Hausa (ha)": "Shawarar mai kula kan wannan takarda",
           "label::English (en)": "7.05 Supervisor decision on this form"},
        relevant="${q7_04_supervisor} != ''")
    row(type="end_group", name="s7_end")

    return R


def choices_rows() -> list[dict]:
    """Choice lists.

    Stored values deliberately avoid the paper form's sentinel numbers where a
    collision existed (defects D1, D2). Where the paper code is safe it is kept,
    so that paper and digital rounds remain comparable.
    """
    C: list[dict] = []

    def opt(list_name, name, ha, en):
        C.append({"list_name": list_name, "name": name,
                  "label::Hausa (ha)": ha, "label::English (en)": en})

    opt("yes_no", "1", "Eh", "Yes")
    opt("yes_no", "2", "A'a", "No")

    opt("yes_no_dk", "1", "Eh", "Yes")
    opt("yes_no_dk", "2", "A'a", "No")
    opt("yes_no_dk", "8", "Ban sani ba", "Do not know")

    opt("consent", "1", "An yarda", "Consent given")
    opt("consent", "2", "An ƙi yarda", "Consent refused")

    opt("sex", "1", "Namiji", "Male")
    opt("sex", "2", "Mace", "Female")

    opt("age_unit", "years", "Shekaru", "Years")
    opt("age_unit", "months", "Watanni", "Months (under 5 years)")

    opt("measured", "measured", "An auna", "Measured")
    opt("measured", "not_measured", "Ba a auna ba", "Not measured")

    opt("measure_position", "1", "A kwance", "Recumbent length")
    opt("measure_position", "2", "A tsaye", "Standing height")

    opt("card_seen", "1", "An ga katin", "Card seen")
    opt("card_seen", "2", "Ba a ga katin ba", "No card seen")

    opt("document_type", "card", "Katin rigakafi", "Vaccination card")
    opt("document_type", "copy", "Kwafin katin", "Card copy")
    opt("document_type", "electronic", "Rikodin lantarki", "Electronic record")

    opt("photo_taken", "1", "Eh", "Yes")
    opt("photo_taken", "2", "A'a, babu shi", "No, not available")
    opt("photo_taken", "3", "Mai kulawa ya ƙi", "Caregiver declined")

    for code, ha, en in [
        ("1", "Ya kammala", "Completed"),
        ("2", "An ƙi", "Refused"),
        ("3", "Babu babban mutum bayan ziyara uku",
         "No competent adult after three visits"),
        ("4", "Gidan babu kowa ko an rushe", "Dwelling vacant or demolished"),
    ]:
        opt("result_of_visit", code, ha, en)

    for code, ha, en in [
        ("1", "Shugaban gida", "Head"), ("2", "Mata/miji", "Spouse"),
        ("3", "Ɗa ko 'ya", "Son or daughter"), ("4", "Iyaye", "Parent"),
        ("5", "Wani dangi", "Other relative"), ("6", "Ba dangi ba", "Not related"),
    ]:
        opt("relationship", code, ha, en)

    # D2 - w-prefixed values so no substantive value equals a sentinel.
    water = ["Piped into dwelling", "Piped into compound", "Public tap or standpipe",
             "Tube well or borehole", "Protected dug well", "Unprotected dug well",
             "Protected spring", "Unprotected spring", "Rainwater",
             "Tanker or cart", "Surface water"]
    water_ha = ["Famfo cikin gida", "Famfo cikin harabar gida", "Famfon jama'a",
                "Rijiyar burtsatse", "Rijiya mai kariya", "Rijiya babu kariya",
                "Marmaro mai kariya", "Marmaro babu kariya", "Ruwan sama",
                "Tanka ko keken ruwa", "Ruwan rafi"]
    for i, (en, ha) in enumerate(zip(water, water_ha), start=1):
        opt("water_source", f"w{i:02d}", ha, en)

    # D1 - t-prefixed for the same reason. Paper code 9 was "No facility or bush"
    # and also the no-answer sentinel.
    toilet = ["Flush to sewer", "Flush to septic tank", "Flush to pit latrine",
              "Ventilated improved pit", "Pit latrine with slab",
              "Pit latrine without slab", "Composting toilet", "Bucket",
              "No facility or bush"]
    toilet_ha = ["Bayan gida mai famfo zuwa bututu", "Bayan gida mai tanki",
                 "Bayan gida mai famfo zuwa rami", "Rami mai iska",
                 "Rami mai murfi", "Rami babu murfi", "Bayan gida na takin",
                 "Guga", "Babu bayan gida / daji"]
    for i, (en, ha) in enumerate(zip(toilet, toilet_ha), start=1):
        opt("toilet", f"t{i:02d}", ha, en)

    opt("handwashing", "1", "An gani, akwai sabulu da ruwa",
        "Observed, soap and water")
    opt("handwashing", "2", "An faɗa kawai, ba a gani ba",
        "Reported only, not observed")
    opt("handwashing", "3", "Babu", "Not present")

    for code, ha, en in [
        ("A", "Rediyo", "Radio"), ("B", "Talabijin", "Television"),
        ("C", "Wayar hannu", "Mobile telephone"), ("D", "Keke", "Bicycle"),
        ("E", "Babur", "Motorcycle"), ("F", "Mota", "Car or truck"),
        ("G", "Firij", "Refrigerator"),
        ("H", "Babu ɗaya daga cikin waɗannan", "None of these"),
    ]:
        opt("assets", code, ha, en)

    for code, ha, en in [
        ("1", "Mai kulawa ya ƙi", "Caregiver refused"),
        ("2", "Yaron ba ya nan", "Child absent"),
        ("3", "Bai iya bayar da shi ba", "Unable to produce"),
        ("4", "Kwanon ya lalace", "Container spoiled"),
        ("96", "Wani dalili", "Other"),
    ]:
        opt("no_specimen_reason", code, ha, en)

    opt("supervisor_decision", "1", "An karɓa", "Accept")
    opt("supervisor_decision", "2", "A gyara", "Return for correction")
    opt("supervisor_decision", "3", "An soke", "Void")

    return C


def settings_rows() -> list[dict]:
    return [{
        "form_title": "Integrated Child Health and AMR Survey 2026 - Household",
        "form_id": FORM_ID,
        "version": FORM_VERSION,
        "default_language": "Hausa (ha)",
        "style": "pages",
        "allow_choice_duplicates": "no",
        # Submissions carry names, GPS to the dwelling entrance, structure
        # numbers and specimen identifiers. Encryption is configured, not
        # optional. The private key is held by the survey manager and is NOT
        # in this repository. See docs/data_protection.md.
        "public_key": "PLACEHOLDER_REPLACE_WITH_REAL_BASE64_PUBLIC_KEY",
    }]


def write_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(SURVEY_COLUMNS)
    for r in survey_rows():
        ws.append([r.get(c, "") for c in SURVEY_COLUMNS])

    ws2 = wb.create_sheet("choices")
    ws2.append(CHOICES_COLUMNS)
    for r in choices_rows():
        ws2.append([r.get(c, "") for c in CHOICES_COLUMNS])

    settings = settings_rows()
    ws3 = wb.create_sheet("settings")
    cols = list(settings[0].keys())
    ws3.append(cols)
    for r in settings:
        ws3.append([r.get(c, "") for c in cols])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def validate(xlsx: Path, xml: Path) -> tuple[bool, str]:
    """Convert with pyxform. A form that does not convert is not a form."""
    result = subprocess.run(
        [sys.executable, "-m", "pyxform.xls2xform", str(xlsx), str(xml)],
        capture_output=True, text=True,
    )
    return result.returncode == 0, (result.stdout + result.stderr).strip()


def main() -> None:
    xlsx = FORM_DIR / f"{FORM_ID}.xlsx"
    xml = FORM_DIR / f"{FORM_ID}.xml"
    log = FORM_DIR / "conversion_log.txt"

    write_workbook(xlsx)
    survey, choices = survey_rows(), choices_rows()
    print(f"  survey rows   {len(survey)}")
    print(f"  choice rows   {len(choices)}")
    print(f"  written       {xlsx.name}")

    ok, output = validate(xlsx, xml)
    header = (f"pyxform conversion of {xlsx.name}\n"
              f"form_id {FORM_ID}  version {FORM_VERSION}\n"
              f"result: {'SUCCESS' if ok else 'FAILURE'}\n"
              + "-" * 60 + "\n")
    log.write_text(header + output + "\n", encoding="utf-8")

    print(f"\n  conversion:   {'SUCCESS' if ok else 'FAILURE'}")
    if output:
        print("  " + output.replace("\n", "\n  ")[:3000])
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
